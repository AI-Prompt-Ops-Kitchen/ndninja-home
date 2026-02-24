#!/usr/bin/env python3
"""
Import SOP content from Evelyn's Excel files into the ROYS database.

Expected Excel structure (per file or per sheet):
- Column A: SOP Code
- Column B: Section Number
- Column C: Section Title
- Column D: Content Body
- Column E: Content Tier (standard/enhanced)
- Column F: Standard Combo Key (e.g., "ISO_13485" or "21_CFR_820+ISO_13485")

Usage:
    python import_content.py --file path/to/content.xlsx [--sheet "Sheet1"] [--dry-run]
"""
import argparse
import asyncio
import sys
import os

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "backend"))

from openpyxl import load_workbook
from sqlalchemy import select
from database import async_session
from models.content import SOP, ContentBlock


async def import_excel(filepath: str, sheet_name: str | None, dry_run: bool):
    wb = load_workbook(filepath, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:
            continue
        sop_code, section_num, section_title, body, tier, combo_key = row[:6]
        if tier not in ("standard", "enhanced"):
            print(f"  WARNING: Row {i} has invalid tier '{tier}', skipping")
            continue
        rows.append({
            "sop_code": str(sop_code).strip(),
            "section_number": str(section_num).strip(),
            "section_title": str(section_title).strip(),
            "body": str(body).strip(),
            "content_tier": str(tier).strip(),
            "combo_key": str(combo_key).strip(),
        })

    print(f"Parsed {len(rows)} content rows from {filepath}")

    if dry_run:
        for r in rows[:5]:
            print(f"  [{r['sop_code']}] {r['section_number']} {r['section_title']} ({r['content_tier']}, {r['combo_key']})")
        if len(rows) > 5:
            print(f"  ... and {len(rows) - 5} more")
        print("Dry run — no changes made.")
        return

    async with async_session() as db:
        created = 0
        skipped = 0
        for r in rows:
            # Find or skip SOP
            result = await db.execute(select(SOP).where(SOP.code == r["sop_code"]))
            sop = result.scalar_one_or_none()
            if not sop:
                print(f"  WARNING: SOP '{r['sop_code']}' not found, skipping row")
                skipped += 1
                continue

            # Check for existing block
            result = await db.execute(
                select(ContentBlock).where(
                    ContentBlock.sop_id == sop.id,
                    ContentBlock.section_number == r["section_number"],
                    ContentBlock.content_tier == r["content_tier"],
                    ContentBlock.combo_key == r["combo_key"],
                )
            )
            if result.scalar_one_or_none():
                skipped += 1
                continue

            block = ContentBlock(
                sop_id=sop.id,
                section_number=r["section_number"],
                section_title=r["section_title"],
                content_tier=r["content_tier"],
                combo_key=r["combo_key"],
                body=r["body"],
                sort_order=int(r["section_number"].split(".")[0]) if r["section_number"][0].isdigit() else 0,
            )
            db.add(block)
            created += 1

        await db.commit()
        print(f"Done: {created} blocks created, {skipped} skipped")


def main():
    parser = argparse.ArgumentParser(description="Import SOP content from Excel")
    parser.add_argument("--file", required=True, help="Path to .xlsx file")
    parser.add_argument("--sheet", default=None, help="Sheet name (default: active sheet)")
    parser.add_argument("--dry-run", action="store_true", help="Parse and validate without inserting")
    args = parser.parse_args()

    if not os.path.exists(args.file):
        print(f"Error: File not found: {args.file}", file=sys.stderr)
        sys.exit(1)

    asyncio.run(import_excel(args.file, args.sheet, args.dry_run))


if __name__ == "__main__":
    main()
